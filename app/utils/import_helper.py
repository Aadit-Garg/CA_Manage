import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models.user import User
from ..models.client import ClientProfile

# Formatting patterns
PAN_REGEX = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$')
GST_REGEX = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$')
PHONE_REGEX = re.compile(r'^\+?[1-9]\d{1,14}$') # E.164 format
EMAIL_REGEX = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')

def generate_excel_template():
    """
    Generate the client import Excel template with formatting and instructions.
    Returns: BytesIO stream of the excel workbook.
    """
    wb = Workbook()
    
    # ── Sheet 1: Instructions ──────────────────────────────────────
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_instr['A1'] = "CA Manage — Bulk Client Import Instructions"
    ws_instr['A1'].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_instr['A1'].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws_instr.merge_cells("A1:D1")
    ws_instr.row_dimensions[1].height = 40
    
    # Instructions Body
    ws_instr.cell(row=2, column=1, value="").font = Font(name="Calibri", size=11, italic=True)
    ws_instr.cell(row=3, column=1, value="Please follow these guidelines strictly to avoid validation errors when importing:").font = Font(name="Calibri", size=11, italic=True)
    
    # Header Row
    for col_idx, col_name in enumerate(["Column Name", "Requirement", "Validation Format / Rule", "Example Value"], start=1):
        cell = ws_instr.cell(row=4, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws_instr.row_dimensions[4].height = 25

    # Re-write instruction rows cleanly
    instr_rows = [
        ["Full Name", "REQUIRED", "Must not be empty. Limit 150 characters.", "Priya Patel"],
        ["Email", "REQUIRED", "Must be a valid email format. Must be unique.", "priya@pateltrading.com"],
        ["Phone", "Optional", "Valid phone format (e.g. +919876543210). Max 20 digits.", "+919988776655"],
        ["Address", "Optional", "Client billing address. Text field.", "123 MG Road"],
        ["City", "Optional", "Text. Max 100 characters.", "Mumbai"],
        ["State", "Optional", "Text. Max 100 characters.", "Maharashtra"],
        ["Pincode", "Optional", "Max 10 characters.", "400001"],
        ["PAN", "Optional", "Must be 10 characters. Format: 5 letters, 4 digits, 1 letter. Unique.", "ABCDE1234F"],
        ["GST", "Optional", "Must be 15 characters. Indian GSTIN format. Unique.", "27ABCDE1234F1Z5"],
        ["Notes", "Optional", "Any internal consultant notes.", "Imported via bulk template."]
    ]
    
    for r_idx, row in enumerate(instr_rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_instr.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1:
                cell.font = Font(name="Calibri", size=11, bold=True)
            if c_idx == 2:
                cell.font = Font(name="Calibri", size=10, bold=(val=="REQUIRED"), color="B91C1C" if val=="REQUIRED" else "475569")
                cell.alignment = Alignment(horizontal="center")
                
    # Auto fit column widths
    for col in ws_instr.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_instr.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ── Sheet 2: Clients Template ─────────────────────────────────
    ws_temp = wb.create_sheet(title="Clients Template")
    ws_temp.views.sheetView[0].showGridLines = True
    
    headers = ["Full Name", "Email", "Phone", "Address", "City", "State", "Pincode", "PAN", "GST", "Notes"]
    
    # Format Header Row
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='medium', color='1B365D')
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_temp.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border
        
    ws_temp.row_dimensions[1].height = 28
    
    # Add Example Row (Row 2)
    example_row = ["Priya Patel", "priya@pateltrading.com", "+919988776655", "123 MG Road", "Mumbai", "Maharashtra", "400001", "ABCDE1234F", "27ABCDE1234F1Z5", "Sample corporate client."]
    for col_idx, val in enumerate(example_row, start=1):
        cell = ws_temp.cell(row=2, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=11, italic=True, color="64748B")
        
    # Auto-fit columns for Template
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws_temp.column_dimensions[col_letter].width = max(len(header) + 5, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def parse_and_validate_excel(file_stream):
    """
    Parse the uploaded Excel sheet, validating every field against DB unique checks and format regulations.
    Returns:
        tuple (valid_rows, invalid_rows)
        - valid_rows: list of dicts representing valid clients to import.
        - invalid_rows: list of dicts representing invalid rows, including a list of specific error messages.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(file_stream, data_only=True)
    if "Clients Template" not in wb.sheetnames:
        raise ValueError("Invalid template format. The sheet 'Clients Template' is missing.")
        
    ws = wb["Clients Template"]
    
    # Ensure header matches
    expected_headers = ["Full Name", "Email", "Phone", "Address", "City", "State", "Pincode", "PAN", "GST", "Notes"]
    row_1 = [cell.value for cell in ws[1]]
    
    for h in expected_headers:
        if h not in row_1:
            raise ValueError(f"Header column '{h}' is missing from the template.")
            
    header_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    
    valid_rows = []
    invalid_rows = []
    
    # Unique trackers inside the uploaded Excel sheet itself to prevent double inserts from the sheet
    sheet_emails = set()
    sheet_pans = set()
    sheet_gsts = set()
    
    # Start iterating from row 2 (which could be the example row).
    # Skip it if it contains exact example values or starts with the Priya Patel template data
    for r_idx in range(2, ws.max_row + 1):
        row_values = {h: ws.cell(row=r_idx, column=header_indices[h]).value for h in expected_headers}
        
        # Check if row is completely empty
        if not any(row_values.values()):
            continue
            
        # Skip the exact example row if they didn't delete it
        if r_idx == 2 and row_values["Email"] == "priya@pateltrading.com":
            continue
            
        errors = []
        full_name = str(row_values["Full Name"] or '').strip()
        email = str(row_values["Email"] or '').strip().lower()
        phone = str(row_values["Phone"] or '').strip() if row_values["Phone"] else None
        address = str(row_values["Address"] or '').strip() if row_values["Address"] else None
        city = str(row_values["City"] or '').strip() if row_values["City"] else None
        state = str(row_values["State"] or '').strip() if row_values["State"] else None
        pincode = str(row_values["Pincode"] or '').strip() if row_values["Pincode"] else None
        pan = str(row_values["PAN"] or '').strip().upper() if row_values["PAN"] else None
        gst = str(row_values["GST"] or '').strip().upper() if row_values["GST"] else None
        notes = str(row_values["Notes"] or '').strip() if row_values["Notes"] else None

        # ── Validations ─────────────────────────────────────────────
        # 1. Required field checks
        if not full_name:
            errors.append("Full Name is required.")
        if not email:
            errors.append("Email is required.")
            
        # 2. Email format & duplicates
        if email:
            if not EMAIL_REGEX.match(email):
                errors.append(f"Invalid email format: '{email}'.")
            elif email in sheet_emails:
                errors.append(f"Duplicate email in this Excel sheet: '{email}'.")
            else:
                sheet_emails.add(email)
                # Check DB for duplicate email
                if User.query.filter_by(email=email).first():
                    errors.append(f"Email '{email}' is already registered in the system.")
                    
        # 3. Phone format
        if phone and not PHONE_REGEX.match(phone):
            errors.append(f"Invalid phone format: '{phone}'. Use e.g. +919876543210.")
            
        # 4. PAN format & duplicates
        if pan:
            if not PAN_REGEX.match(pan):
                errors.append(f"Invalid PAN format: '{pan}'. Must be 5 letters, 4 digits, 1 letter.")
            elif pan in sheet_pans:
                errors.append(f"Duplicate PAN in this Excel sheet: '{pan}'.")
            else:
                sheet_pans.add(pan)
                # Check DB
                if ClientProfile.query.filter_by(PAN=pan).first():
                    errors.append(f"PAN '{pan}' is already registered in the system.")
                    
        # 5. GST format & duplicates
        if gst:
            if not GST_REGEX.match(gst):
                errors.append(f"Invalid GSTIN format: '{gst}'. E.g. 27ABCDE1234F1Z5")
            elif gst in sheet_gsts:
                errors.append(f"Duplicate GSTIN in this Excel sheet: '{gst}'.")
            else:
                sheet_gsts.add(gst)
                # Check DB
                if ClientProfile.query.filter_by(GST=gst).first():
                    errors.append(f"GSTIN '{gst}' is already registered in the system.")
                    
        client_data = {
            'row_num': r_idx,
            'full_name': full_name,
            'email': email,
            'phone': phone,
            'address': address,
            'city': city,
            'state': state,
            'pincode': pincode,
            'PAN': pan,
            'GST': gst,
            'notes': notes
        }
        
        if errors:
            client_data['errors'] = errors
            invalid_rows.append(client_data)
        else:
            valid_rows.append(client_data)
            
    return valid_rows, invalid_rows
