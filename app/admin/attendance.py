"""
Admin Attendance Management Routes
"""
from datetime import datetime, timezone, timedelta
from flask import render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import current_user
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..extensions import db
from ..models.attendance import Attendance
from ..models.employee import Employee
from ..models.user import User
from . import admin_bp
from ..admin.routes import admin_required
from ..utils.logging import get_logger, log_user_action

logger = get_logger(__name__)


@admin_bp.route('/attendance')
@admin_required
def attendance_dashboard():
    """
    Admin attendance overview — daily view with filter options.
    """
    date_str = request.args.get('date')
    if date_str:
        try:
            view_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            view_date = datetime.now(timezone.utc).date()
    else:
        view_date = datetime.now(timezone.utc).date()

    # Get all employees
    employees = Employee.query.join(User).filter(User.is_active == True).all()

    # Get attendance records for the selected date
    records = Attendance.query.filter_by(date=view_date).all()
    records_by_employee = {r.employee_id: r for r in records}

    # Build attendance rows
    attendance_rows = []
    present_count = 0
    absent_count = 0
    working_count = 0

    for emp in employees:
        record = records_by_employee.get(emp.id)
        row = {
            'employee': emp,
            'record': record,
        }
        if record:
            if record.punch_out_time:
                present_count += 1
            elif record.punch_in_time:
                working_count += 1
        else:
            absent_count += 1

        attendance_rows.append(row)

    stats = {
        'total': len(employees),
        'present': present_count,
        'working': working_count,
        'absent': absent_count,
    }

    return render_template(
        'admin/attendance/dashboard.html',
        attendance_rows=attendance_rows,
        stats=stats,
        view_date=view_date,
    )


@admin_bp.route('/attendance/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_attendance(id):
    """
    Allows admin to correct an attendance record.
    """
    record = Attendance.query.get_or_404(id)
    employee = Employee.query.filter_by(user_id=record.employee_id).first()

    if request.method == 'POST':
        punch_in_str = request.form.get('punch_in_time')
        punch_out_str = request.form.get('punch_out_time')
        status = request.form.get('status')
        notes = request.form.get('notes', '')

        correction_log = {
            'corrected_by': current_user.id,
            'corrected_at': datetime.now(timezone.utc).isoformat(),
            'previous_punch_in': record.punch_in_time.isoformat() if record.punch_in_time else None,
            'previous_punch_out': record.punch_out_time.isoformat() if record.punch_out_time else None,
            'previous_status': record.status,
        }

        if punch_in_str:
            try:
                record.punch_in_time = datetime.strptime(punch_in_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                pass
        if punch_out_str:
            try:
                record.punch_out_time = datetime.strptime(punch_out_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                pass

        if record.punch_in_time and record.punch_out_time:
            td = record.punch_out_time - record.punch_in_time
            record.total_hours = round(td.total_seconds() / 3600.0, 2)

        if status:
            record.status = status

        record.notes = notes
        record.correction_log = str(correction_log)

        db.session.commit()
        
        # Notify the employee about the correction
        from ..utils.notification import create_notification
        create_notification(
            record.employee_id, 
            f"Your attendance for {record.date.strftime('%Y-%m-%d')} has been corrected by {current_user.full_name}.", 
            link=url_for('employee.attendance_history'),
            title='Attendance Corrected',
            category='attendance',
            priority='normal',
            entity_type='Attendance',
            entity_id=record.id
        )

        log_user_action(logger, current_user, 'correct_attendance', module='attendance', entity_type='Attendance', entity_id=id, description=f'Corrected attendance record {id}')
        flash('Attendance record corrected successfully.', 'success')
        return redirect(url_for('admin.attendance_dashboard', date=record.date.strftime('%Y-%m-%d')))

    return render_template('admin/attendance/edit.html', record=record, employee=employee)


@admin_bp.route('/attendance/create', methods=['GET', 'POST'])
@admin_required
def create_attendance():
    """
    Allows admin to manually create an attendance record for an employee.
    """
    employees = Employee.query.join(User).filter(User.is_active == True).all()

    if request.method == 'POST':
        employee_id = request.form.get('employee_id', type=int)
        date_str = request.form.get('date')
        punch_in_str = request.form.get('punch_in_time')
        punch_out_str = request.form.get('punch_out_time')
        status = request.form.get('status', 'Present')
        notes = request.form.get('notes', '')

        try:
            att_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid date.', 'danger')
            return redirect(url_for('admin.create_attendance'))

        # Check if record already exists
        existing = Attendance.query.filter_by(employee_id=employee_id, date=att_date).first()
        if existing:
            flash('Attendance record already exists for this employee on this date. Please edit it instead.', 'warning')
            return redirect(url_for('admin.attendance_dashboard', date=att_date.strftime('%Y-%m-%d')))

        record = Attendance(employee_id=employee_id, date=att_date, status=status, notes=notes)

        if punch_in_str:
            try:
                record.punch_in_time = datetime.strptime(punch_in_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                pass
        if punch_out_str:
            try:
                record.punch_out_time = datetime.strptime(punch_out_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                pass

        if record.punch_in_time and record.punch_out_time:
            td = record.punch_out_time - record.punch_in_time
            record.total_hours = round(td.total_seconds() / 3600.0, 2)

        record.correction_log = f"Manually created by Admin {current_user.id} at {datetime.now(timezone.utc).isoformat()}"

        db.session.add(record)
        db.session.commit()

        log_user_action(logger, current_user, 'create_attendance', module='attendance', entity_type='Attendance', entity_id=record.id, description=f'Created attendance for employee {employee_id} on {date_str}')
        flash('Attendance record created successfully.', 'success')
        return redirect(url_for('admin.attendance_dashboard', date=att_date.strftime('%Y-%m-%d')))

    return render_template('admin/attendance/create.html', employees=employees)


@admin_bp.route('/attendance/export')
@admin_required
def export_attendance():
    """
    Exports attendance records for a specific month or date range to Excel.
    """
    month_str = request.args.get('month')
    if month_str:
        try:
            view_month = datetime.strptime(month_str, '%Y-%m').date()
        except ValueError:
            view_month = datetime.now(timezone.utc).date().replace(day=1)
    else:
        view_month = datetime.now(timezone.utc).date().replace(day=1)

    # Calculate start and end of month
    start_date = view_month
    if view_month.month == 12:
        end_date = view_month.replace(year=view_month.year + 1, month=1) - timedelta(days=1)
    else:
        end_date = view_month.replace(month=view_month.month + 1) - timedelta(days=1)

    records = Attendance.query.filter(Attendance.date >= start_date, Attendance.date <= end_date).order_by(Attendance.date.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {view_month.strftime('%b %Y')}"

    # Define styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ['Date', 'Employee Code', 'Employee Name', 'Status', 'Punch In', 'Punch Out', 'Total Hours', 'Location', 'Notes']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for r in records:
        employee_code = r.employee.client_code if hasattr(r.employee, 'client_code') else f"EMP-{r.employee.id}"
        employee_name = r.employee.user.full_name
        punch_in = r.punch_in_time.strftime('%I:%M %p') if r.punch_in_time else '—'
        punch_out = r.punch_out_time.strftime('%I:%M %p') if r.punch_out_time else '—'
        hours = str(r.total_hours) if r.total_hours is not None else '0'

        ws.append([
            r.date.strftime('%Y-%m-%d'),
            employee_code,
            employee_name,
            r.status,
            punch_in,
            punch_out,
            hours,
            r.location_status or '—',
            r.notes or ''
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Attendance_Export_{view_month.strftime('%b_%Y')}.xlsx"
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
